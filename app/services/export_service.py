"""
Generacion de reportes en PDF (reportlab) y Excel (openpyxl) en memoria.

Recibe la forma tabular uniforme (titulo, columnas, filas) y devuelve bytes.
El router los envuelve en StreamingResponse. No usa pandas.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _fecha_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _celda(valor) -> str:
    return "" if valor is None else str(valor)


def to_excel(titulo: str, columnas: list, filas: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append([titulo])
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws.append([f"Generado: {_fecha_str()}"])
    ws["A2"].font = Font(size=9, italic=True, color="7F7F7F")
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(list(columnas))
    relleno = PatternFill("solid", fgColor="4A90E2")
    for col_idx in range(1, len(columnas) + 1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = relleno
        c.alignment = Alignment(horizontal="left")

    if filas:
        for fila in filas:
            ws.append([_celda(fila.get(col)) for col in columnas])
    else:
        ws.append(["Sin datos para el periodo seleccionado"])

    for col_idx, col in enumerate(columnas, start=1):
        anchos = [len(str(col))] + [len(_celda(f.get(col))) for f in filas]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max(anchos) + 2, 12), 50)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf(titulo: str, columnas: list, filas: list) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=titulo,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(titulo, estilos["Title"]),
        Paragraph(f"Generado: {_fecha_str()}", estilos["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    if filas:
        tabla_data = [list(columnas)] + [
            [_celda(f.get(col)) for col in columnas] for f in filas
        ]
        tabla = Table(tabla_data, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A90E2")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elementos.append(tabla)
    else:
        elementos.append(Paragraph("Sin datos para el periodo seleccionado.", estilos["Normal"]))

    doc.build(elementos)
    return buffer.getvalue()
